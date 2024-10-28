import openpyxl


def get_data_for_load(xlsx_file_path: str):
    

    # Give the location of the file
    path = xlsx_file_path

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    row = sheet_obj.max_row
    rez = {}
    for i in range(1, row + 1):
        art = sheet_obj.cell(row=i, column=1).value
        links = [linc for 
                 linc in 
                 sheet_obj.cell(row=i, column=2).value.split(',')]
        rez[art] = links

    
    return rez
        
if __name__ == '__main__':
    rez: dict = get_data_for_load('sours.xlsx')
    for item in rez.items():
        print(item)